import argparse
import json
import os
from copy import deepcopy
from datetime import datetime

from openpyxl import load_workbook, Workbook


def parse_statement(filename, **kwargs):
    first_data_row = kwargs.get('first_data_row', 9)
    sheet_name = kwargs.get('sheet_name', 'BGPMainScreen')
    # Cuenta	Tarjeta  	Fecha transacción	Fecha proceso	Descripción
    # Referencia	Categoría	Cargos (Db)	Pagos (Cr)

    mappings = {
        'account': {'name': 'Cuenta', 'col': 1},
        'card': {'name': 'Tarjeta', 'col': 2},
        'transaction_date': {'name': 'Fecha transacción', 'col': 3},
        'process_date': {'name': 'Fecha proceso', 'col': 4},
        'description': {'name': 'Descripcion', 'col': 5},
        'reference': {'name': 'Referencia', 'col': 8},
        'category': {'name': 'Categoría', 'col': 9},
        'charges': {'name': 'Cargos', 'col': 10},
        'payments': {'name': 'Pagos', 'col': 11},
    }

    wb = load_workbook(filename=filename, data_only=True)
    sheet = wb[sheet_name]
    row_num = 1
    row_data_list = list()
    for row in sheet.rows:
        row_dict = dict()
        if row_num >= first_data_row:
            try:
                # charges = row[mappings['charges']['col']].value

                for attribute_name in mappings.keys():
                    row_dict[attribute_name] = mappings[attribute_name]
                    row_dict[attribute_name]['value'] = row[mappings[attribute_name]['col']].value
                    row_dict[attribute_name]['row_num'] = row_num
                    row_dict[attribute_name]['source_file'] = filename
                # print(f'{row_num}. {charges} {row_dict["charges"]["value"]} {type(charges)}')
                row_data_list.append(deepcopy(row_dict))
            except Exception as e:
                raise e
        # if row_num >= first_data_row:
        #     print(f'{row_num}.  {row_data_list[row_num - first_data_row]["charges"]["value"]}')
        # print('-' * 60)
        row_num += 1
    # print('86-' * 60)
    # row_num = 86
    # print(f'{row_num}.  {row_data_list[row_num - first_data_row]["charges"]["value"]}')value
    return row_data_list


def write_to_excel(filename, data_list, **kwargs):
    wb = Workbook()
    sheet = wb.create_sheet()
    row = 1
    col = 1
    for attribute_name in data_list[0].keys():
        sheet.cell(row=row, column=col, value=data_list[0][attribute_name]['name'])
        col += 1
    row += 1
    for data_dict in data_list:
        col = 1
        for attribute_name in data_dict.keys():
            sheet.cell(row=row, column=col, value=data_dict[attribute_name]['value'])
            col += 1
        row += 1

    wb.save(filename)


def scan_folder(folder, **kwargs):
    result_list = list()
    file_count = 0
    for dir_path, dir_names, file_names in os.walk(folder, followlinks=False):
        for file_name in [f for f in file_names if f.endswith(".xlsx")]:
            excel_filename = os.path.normpath(os.path.join(dir_path, file_name))
            if kwargs.get('verbose'):
                print(f'Processing {file_name}......')
            # print(os.path.normpath(os.path.join(dir_path, file_name)))
            file_result = parse_statement(excel_filename, **kwargs)
            if kwargs.get('verbose'):
                print(f' Processed {len(file_result)} records')
            if len(file_result) > 0:
                result_list += file_result
                file_count += 1
    return result_list, file_count


def ensure_dir(file_path):
    directory = os.path.dirname(file_path)
    if not os.path.exists(directory):
        os.makedirs(directory)


if __name__ == '__main__':
    # Create the parser
    my_parser = argparse.ArgumentParser(description='Convert statements to one single Excel file')
    my_parser.add_argument('-v',
                           '--verbose',
                           action='store_true',
                           help='Prints extra info')
    my_parser.add_argument('-j',
                           '--json',
                           action='store_true',
                           help='Saves file to json format in the Output folder')
    my_parser.add_argument('-o',
                           '--output-folder',
                           action='store',
                           help='Output folder',
                           default='./output')
    my_parser.add_argument('-d',
                          '--data-folder',
                          action='store',
                          help='Data folder. Place your Excel statements herw',
                          default='./data')

    # Add the arguments
    # my_parser.add_argument('Path',
    #                        metavar='path',
    #                        type=str,
    #                        help='the path to list')
    args = my_parser.parse_args()
    export_json = args.json
    verbose = args.verbose
    ensure_dir(args.output_folder)

    row_list, files_processed = scan_folder(args.data_folder, verbose=verbose)

    # print(row_list)
    if len(row_list) > 0:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if export_json:
            json_output_file = os.path.join(args.output_folder, f'{timestamp}_estado_cuenta.json')
            with open(json_output_file, 'w', encoding='utf-8') as json_file:
                json_file.write(json.dumps(row_list))
        excel_output_file = os.path.join(args.output_folder, f'{timestamp}_estado_cuenta.xlsx')
        write_to_excel(excel_output_file, row_list)

        msg = f' Wrote {len(row_list)} records from {files_processed} files in {excel_output_file}'
        print('-' * (len(msg) + 20))
        print(msg)
        print('-' * (len(msg) + 20))
    else:
        print(f'Found no excel files in {args.data_folder}')
